import argparse
import ast
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import time
import requests
from openpyxl import Workbook

from agent.base_agent import APIConfig
from canonicalize import normalize_to_canonical
from prompt.match_item import PROMPT_MATCH_ITEM
from utils import *


def _gold_sort_key(gold_id: str):
    """Numeric-aware, deterministic ordering key for gold ids.

    Numeric ids sort before string ids and by integer value (so '2' < '10',
    NOT lexical '10' < '2' which is wrong for ids reaching 18). Non-numeric ids
    (e.g. 'EX-NN', 'FT-NN') sort after numerics, lexically among themselves.
    """
    s = str(gold_id)
    if s.isdigit():
        return (0, int(s), "")
    return (1, 0, s)


def aggregate_ballots(
    ballots: List[List[Tuple[str, Optional[str]]]],
    pred_order: Optional[List[str]] = None,
) -> List[Tuple[str, Optional[str]]]:
    """Aggregate K matcher ballots into one (pred_id, gold_id|None) per pred.

    PURE / network-free / total over its inputs. Union semantics (tau=1):

    - Each ballot is a list of (pred_id, gold_id|None) pairs. ``None`` is the
      ABSENCE of a vote, not a competing vote.
    - For each predicted id we collect only the non-None gold votes across all
      ballots (str-coerced; int 3 and str '3' collapse to one '3'). If any
      non-None gold vote exists, the result is the MODE; ties broken by the
      SMALLEST gold id via :func:`_gold_sort_key` (numeric-aware, deterministic
      and ballot-order-independent). A pred with ZERO non-None votes emits
      ``(pred, None)``.
    - Within a single ballot a pred is deduped (counts once per gold) so one
      flaky ballot repeating a pred cannot outvote other ballots.
    - Malformed rows (not exactly 2 elements, or a falsy/None pred_id) are
      skipped defensively so a stray row cannot crash the record.
    - Output: exactly ONE row per distinct pred_id. Order follows ``pred_order``
      when supplied (each listed pred appears once; omitted-by-all -> None);
      otherwise first-appearance across the concatenated ballots.

    Args:
        ballots: surviving (non-None) per-run ballots. May include valid-empty
            ``[]`` ballots (they contribute nothing but do not veto).
        pred_order: canonical predicted-id order (e.g. ``list(pred_items)``).

    Returns:
        ``[(pred_id, gold_id|None), ...]`` — empty list for empty/no-vote input.
    """
    # Per pred: set of (gold) seen within the CURRENT ballot (for intra-ballot
    # dedupe) is tracked per ballot; cross-ballot we accumulate a Counter.
    vote_counts: Dict[str, "Counter"] = defaultdict(Counter)
    # Track first-appearance order of pred ids across concatenated ballots.
    appearance_order: List[str] = []
    seen_preds = set()

    def _note_pred(pred_id: str) -> None:
        if pred_id not in seen_preds:
            seen_preds.add(pred_id)
            appearance_order.append(pred_id)

    for ballot in ballots or []:
        if not ballot:
            continue
        intra_seen: Dict[str, set] = defaultdict(set)
        for row in ballot:
            # Defend against malformed rows (1-/3-tuples, None pred).
            try:
                if len(row) != 2:
                    continue
            except TypeError:
                continue
            pred_id, gold_id = row[0], row[1]
            if pred_id is None:
                continue
            _note_pred(pred_id)
            if gold_id is None:
                # None is absence of a vote; record the pred but cast no gold vote.
                continue
            gold_key = str(gold_id)
            # Intra-ballot dedupe: a (pred, gold) pair counts once per ballot.
            if gold_key in intra_seen[pred_id]:
                continue
            intra_seen[pred_id].add(gold_key)
            vote_counts[pred_id][gold_key] += 1

    order = pred_order if pred_order is not None else appearance_order

    result: List[Tuple[str, Optional[str]]] = []
    emitted = set()
    for pred_id in order:
        if pred_id in emitted:
            continue
        emitted.add(pred_id)
        counts = vote_counts.get(pred_id)
        if not counts:
            # Zero non-None gold votes (or pred omitted by all ballots) -> None.
            result.append((pred_id, None))
            continue
        max_count = max(counts.values())
        winners = [g for g, c in counts.items() if c == max_count]
        gold = min(winners, key=_gold_sort_key)
        result.append((pred_id, gold))

    return result


class ScoringPipeline:
    """
    Scorer for WebProberBench agent outputs.

    Responsibilities:
    1. Load dataset records and corresponding agent outputs.
    2. Parse gold checklist items and predicted checklist items.
    3. Match predicted items to gold items via an LLM.
    4. Compute precision, recall, F1, and checklist coverage.
    5. Aggregate metrics and export JSON/Excel reports.
    """

    def __init__(
        self,
        dataset_path: Path,
        output_root: Path,
        api_config: APIConfig,
        version: str,
        use_checklist_fallback: bool = False,
        canonicalize: bool = True,
        match_votes: int = 1,
    ) -> None:
        self.dataset_path = dataset_path
        self.output_root = output_root
        self.version = version
        self.api_config = api_config
        self.use_checklist_fallback = use_checklist_fallback
        self.canonicalize = canonicalize
        # Number of matcher ballots to union (D2). Coerce defensively to int >= 1:
        # 0/negative would make range(K) empty -> zero ballots -> every record
        # falsely becomes empty_match. K=1 == exact single-call behavior.
        self.match_votes = max(1, int(match_votes))
        self.dataset = self._load_dataset()

        # Task IDs where result files are missing.
        self.missing_result_ids: List[str] = []
        # Task IDs where matching returned no usable links.
        self.empty_match_ids: List[str] = []
        # Task IDs where result_extracted.md exists but no items could be parsed
        # (detection output format mismatch). These are flagged, NOT silently
        # treated as an all-pass checklist (which would fabricate recall=0).
        self.parse_error_ids: List[str] = []

        # Stable ordering used when rendering grouped statistics.
        self.ordered_categories = [
            "Presentation", "Search", "Tool", "Commerce",
            "Data Management", "Workflow", "User-Generated Content",
        ]
        self.ordered_classes = ["functionality", "constraint", "interaction", "content"]

    def run(self) -> None:
        """
        Execute the full scoring workflow:
        1. Iterate through all dataset records.
        2. Score each record independently.
        3. Aggregate metrics and write final reports.
        """
        # Clear stale missing-result logs from prior runs.
        missing_path = self.output_root / "missing_results.json"
        if missing_path.exists():
            missing_path.unlink()

        # Initialize in-memory accumulators.
        aggregators = self._initialize_aggregators()
        
        # Score each record and fold results into accumulators.
        for record_id, record in self.dataset.items():
            aggregators['total_count'] += 1
            
            output_dir = self.output_root / record_id
            if not output_dir.exists():
                print(f"Warning: output dir not found for {record_id}, skipping.")
                continue
                
            print_green(f"Scoring {record_id}")
            
            # Track how many gold items belong to each class.
            self._update_class_item_counts(record, aggregators['class_item_counts'])
            
            # Process one record and obtain per-record metrics.
            result_bundle = self._process_record(record_id, record, output_dir)
            if result_bundle is None:
                continue
                
            # Merge per-record output into global aggregates.
            self._update_aggregators(result_bundle, record, aggregators)

        # Persist missing-result task IDs for debugging/auditing.
        if self.missing_result_ids:
            self._write_json(missing_path, {"missing_result": self.missing_result_ids})

        # Compute averages and write final artifacts.
        self._compute_and_save_final_results(aggregators)

    def _initialize_aggregators(self) -> Dict:
        """Create and return all aggregate counters used across records."""
        return {
            'total_count': 0,
            'scored_count': 0,
            'scored_count_no_missing': 0,
            'total_metrics': {"precision": 0.0, "recall": 0.0, "f1": 0.0, "coverage": 0.0},
            'total_metrics_no_missing': {"precision": 0.0, "recall": 0.0, "f1": 0.0, "coverage": 0.0},
            'total_num_pred_item': 0,
            'total_num_pred_item_no_missing': 0,
            'category_totals': {},
            'category_counts': {},
            'class_totals': {},
            'class_counts': {},
            'class_item_counts': {cls: 0 for cls in self.ordered_classes},
            'class_coverage_totals': {cls: 0.0 for cls in self.ordered_classes},
            'class_coverage_counts': {cls: 0 for cls in self.ordered_classes},
        }

    def _update_class_item_counts(self, record: dict, class_item_counts: Dict[str, int]) -> None:
        """Count gold checklist items per class for denominator reporting."""
        gold_items = self._parse_gold_checklist(record)
        for item in gold_items.values():
            cls = item.get("class")
            if cls in class_item_counts:
                class_item_counts[cls] += 1

    def _update_aggregators(self, result_bundle: dict, record: dict, aggregators: Dict) -> None:
        """Update all global aggregates using one processed record bundle."""
        metrics = result_bundle.get("overall")
        class_metrics = result_bundle.get("by_class", {})
        match_ids = result_bundle.get("match_ids")
        gold_items = result_bundle.get("gold_items", {})
        missing_result = metrics.get("missing_result", False)
        empty_match = metrics.get("empty_match", False)
        parse_error = metrics.get("parse_error", False)

        # Update overall metrics (includes missing/empty/parse-error cases).
        self._accumulate_metrics(aggregators['total_metrics'], metrics)
        aggregators['scored_count'] += 1
        aggregators['total_num_pred_item'] += int(metrics.get("num_pred_item", 0))

        # Update "no_missing" slice (exclude missing/empty/parse-error records).
        if (not missing_result) and (not empty_match) and (not parse_error):
            self._accumulate_metrics(aggregators['total_metrics_no_missing'], metrics)
            aggregators['scored_count_no_missing'] += 1
            aggregators['total_num_pred_item_no_missing'] += int(metrics.get("num_pred_item", 0))

        # Update category-level aggregate metrics.
        self._update_category_stats(record, metrics, aggregators)
        
        # Update class-level aggregate metrics.
        self._update_class_stats(class_metrics, aggregators)
        
        # Update class-level coverage averages.
        self._update_class_coverage(match_ids, gold_items, aggregators)

    def _accumulate_metrics(self, target: Dict[str, float], source: Dict[str, float]) -> None:
        """Add a metric dictionary into an accumulator dictionary in place."""
        for key in target:
            target[key] += float(source.get(key, 0.0))

    def _update_category_stats(self, record: dict, metrics: Dict, aggregators: Dict) -> None:
        """Accumulate metrics grouped by high-level task category."""
        category = record.get("category", "Unknown")
        
        if category not in aggregators['category_totals']:
            aggregators['category_totals'][category] = {k: 0.0 for k in aggregators['total_metrics']}
            aggregators['category_counts'][category] = 0
        
        self._accumulate_metrics(aggregators['category_totals'][category], metrics)
        aggregators['category_counts'][category] += 1

    def _update_class_stats(self, class_metrics: Dict, aggregators: Dict) -> None:
        """Accumulate metrics grouped by checklist class."""
        for cls in self.ordered_classes:
            cls_metrics = class_metrics.get(cls)
            if not cls_metrics or cls_metrics.get("precision") is None:
                continue
                
            if cls not in aggregators['class_totals']:
                aggregators['class_totals'][cls] = {k: 0.0 for k in aggregators['total_metrics']}
                aggregators['class_counts'][cls] = 0
            
            self._accumulate_metrics(aggregators['class_totals'][cls], cls_metrics)
            aggregators['class_counts'][cls] += 1

    def _update_class_coverage(
        self, 
        match_ids: List[Tuple[str, Optional[str]]], 
        gold_items: Dict[str, dict],
        aggregators: Dict
    ) -> None:
        """Accumulate per-record class coverage; average only where class exists."""
        if not gold_items:
            return

        matched_gold_ids = {gold_id for _, gold_id in (match_ids or []) if gold_id is not None}
        class_total_in_record: Dict[str, int] = {cls: 0 for cls in self.ordered_classes}
        class_covered_in_record: Dict[str, int] = {cls: 0 for cls in self.ordered_classes}

        for gold_id, gold_meta in gold_items.items():
            cls = gold_meta.get("class")
            if cls not in class_total_in_record:
                continue
            class_total_in_record[cls] += 1
            if gold_id in matched_gold_ids:
                class_covered_in_record[cls] += 1

        for cls in self.ordered_classes:
            total_cls = class_total_in_record[cls]
            if total_cls == 0:
                continue
            coverage = class_covered_in_record[cls] / total_cls
            aggregators['class_coverage_totals'][cls] += coverage
            aggregators['class_coverage_counts'][cls] += 1

    def _compute_and_save_final_results(self, aggregators: Dict) -> None:
        """Compute final averages from aggregates and persist reports."""
        scored_count = aggregators['scored_count']
        scored_count_no_missing = aggregators['scored_count_no_missing']
        
        # Compute overall means.
        avg_metrics = self._compute_average_metrics(
            aggregators['total_metrics'], scored_count
        )
        avg_metrics_no_missing = self._compute_average_metrics(
            aggregators['total_metrics_no_missing'], scored_count_no_missing
        )
        
        # Compute category-level means.
        category_avg = self._compute_category_averages(aggregators)
        
        # Compute class-level means.
        class_avg = self._compute_class_averages(aggregators)
        
        # Build unified summary payload.
        merged_avg = {
            "overall": avg_metrics,
            "overall_no_missing": avg_metrics_no_missing,
            "avg_num_pred_item": round(aggregators['total_num_pred_item'] / scored_count, 4) if scored_count else None,
            "by_category": category_avg,
            "by_class": class_avg,
            "counts": {
                "total": aggregators['total_count'],
                "scored": scored_count,
                "missing_result": len(self.missing_result_ids),
                "empty_match": len(self.empty_match_ids),
                "parse_error": len(self.parse_error_ids),
                "scored_no_missing": scored_count_no_missing,
            },
        }
        
        # Write output files.
        self._write_json(self.output_root / "score_avg.json", merged_avg)
        self._write_category_excel(category_avg, class_avg, avg_metrics)
        
        # Print top-line summary for quick inspection.
        if avg_metrics:
            print_green(
                f"Overall average score (P/R/F1): "
                f"{avg_metrics.get('precision', 0.0):.4f}/"
                f"{avg_metrics.get('recall', 0.0):.4f}/"
                f"{avg_metrics.get('f1', 0.0):.4f}"
            )
        if self.empty_match_ids:
            print_red(f"Empty matches for: {self.empty_match_ids}")
        if self.parse_error_ids:
            print_red(
                f"Parse errors (result_extracted.md present but unparseable, "
                f"flagged NOT all-pass) for: {self.parse_error_ids}"
            )
            self._write_json(self.output_root / "parse_errors.json",
                             {"parse_error": self.parse_error_ids})

    def _compute_average_metrics(
        self, 
        totals: Dict[str, float], 
        count: int
    ) -> Optional[Dict[str, float]]:
        """Return rounded averages for a metric-total map, or None when empty."""
        if not count:
            return None
        return {
            **{k: round(v / count, 4) for k, v in totals.items()},
            "count": count
        }

    def _compute_category_averages(self, aggregators: Dict) -> Dict[str, dict]:
        """Compute per-category averaged metrics in a stable display order."""
        category_avg = {}
        
        # Emit known categories first for deterministic tables.
        for category in self.ordered_categories:
            if category not in aggregators['category_totals']:
                continue
            count = aggregators['category_counts'].get(category, 0)
            if not count:
                continue
            totals = aggregators['category_totals'][category]
            category_avg[category] = {
                **{k: round(v / count, 4) for k, v in totals.items()},
                "count": count,
            }
        
        # Append any extra categories not present in the predefined list.
        for category, totals in aggregators['category_totals'].items():
            if category in category_avg:
                continue
            count = aggregators['category_counts'].get(category, 0)
            if not count:
                continue
            category_avg[category] = {
                **{k: round(v / count, 4) for k, v in totals.items()},
                "count": count,
            }
        
        return category_avg

    def _compute_class_averages(self, aggregators: Dict) -> Dict[str, dict]:
        """Compute per-class averaged metrics and coverage."""
        class_avg = {}
        
        for cls in self.ordered_classes:
            total_cls = aggregators['class_item_counts'].get(cls, 0)
            count = aggregators['class_counts'].get(cls, 0)
            
            # Average class metrics over records where that class was scored.
            avg_metrics_cls = {}
            if count and cls in aggregators['class_totals']:
                avg_metrics_cls = {
                    k: round(v / count, 4) 
                    for k, v in aggregators['class_totals'][cls].items()
                }
            
            # Coverage is averaged over records that contain this class.
            coverage = 0.0
            coverage_count = aggregators['class_coverage_counts'].get(cls, 0)
            if coverage_count:
                coverage = round(
                    aggregators['class_coverage_totals'].get(cls, 0.0) / coverage_count, 4
                )
            
            class_avg[cls] = {
                **avg_metrics_cls,
                "count": total_cls,
                "coverage": coverage,
            }
        
        return class_avg

    def _write_category_excel(
        self,
        category_avg: Dict[str, dict],
        class_avg: Dict[str, dict],
        avg_metrics: Optional[Dict[str, float]],
    ) -> None:
        """Write an Excel report with category/class and overall score views."""
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "category_scores"

        # Create header rows.
        self._create_excel_header(sheet)
        
        # Fill data rows.
        self._fill_excel_data(sheet, category_avg, class_avg, avg_metrics)
        
        # Persist workbook.
        output_path = self.output_root / f"{self.version}_score.xlsx"
        workbook.save(output_path)

    def _create_excel_header(self, sheet) -> None:
        """Create Excel header columns and row labels."""
        # Keep first cell blank for row labels.
        sheet.cell(row=1, column=1, value="")
        
        # Category columns.
        for idx, category in enumerate(self.ordered_categories, start=2):
            sheet.cell(row=1, column=idx, value=category)
        
        # Category overall column.
        category_overall_col = len(self.ordered_categories) + 2
        sheet.cell(row=1, column=category_overall_col, value="Overall")
        
        # Class columns.
        class_start_col = category_overall_col + 1
        for offset, cls in enumerate(self.ordered_classes):
            sheet.cell(row=1, column=class_start_col + offset, value=cls)
        
        # Class overall column.
        class_overall_col = class_start_col + len(self.ordered_classes)
        sheet.cell(row=1, column=class_overall_col, value="Overall")

        # Additional CPRF columns: Coverage/Precision/Recall/F1 for each class + overall.
        cprf_start_col = class_overall_col + 1
        for offset, cls in enumerate(self.ordered_classes):
            sheet.cell(row=1, column=cprf_start_col + offset, value=f"{cls}_CPRF")
        cprf_overall_col = cprf_start_col + len(self.ordered_classes)
        sheet.cell(row=1, column=cprf_overall_col, value="Overall_CPRF")
        
        # Row labels.
        score_row_label = f"{self.version}_score"
        coverage_row_label = f"{self.version}_coverage"
        sheet.cell(row=2, column=1, value=score_row_label)
        sheet.cell(row=3, column=1, value=coverage_row_label)

    def _fill_excel_data(
        self, 
        sheet, 
        category_avg: Dict[str, dict],
        class_avg: Dict[str, dict],
        avg_metrics: Optional[Dict[str, float]]
    ) -> None:
        """Populate Excel score/coverage rows."""
        def pct_str(value: float) -> str:
            """Format decimal metric as a percentage string with one decimal place."""
            return f"{round(value * 100, 1):.1f}"

        # Fill category columns.
        for idx, category in enumerate(self.ordered_categories, start=2):
            metrics = category_avg.get(category)
            if not metrics:
                continue
            self._fill_metric_cell(sheet, 2, idx, metrics, pct_str)
            sheet.cell(row=3, column=idx, value=pct_str(metrics.get("coverage", 0.0)))

        # Fill class columns.
        class_start_col = len(self.ordered_categories) + 3
        for offset, cls in enumerate(self.ordered_classes):
            metrics = class_avg.get(cls)
            col = class_start_col + offset
            if not metrics:
                continue
            self._fill_metric_cell(sheet, 2, col, metrics, pct_str)
            sheet.cell(row=3, column=col, value=pct_str(metrics.get("coverage", 0.0)))

        # Fill overall columns.
        if avg_metrics:
            category_overall_col = len(self.ordered_categories) + 2
            class_overall_col = class_start_col + len(self.ordered_classes)
            
            self._fill_metric_cell(sheet, 2, category_overall_col, avg_metrics, pct_str)
            sheet.cell(row=3, column=category_overall_col, value=pct_str(avg_metrics.get("coverage", 0.0)))
            
            self._fill_metric_cell(sheet, 2, class_overall_col, avg_metrics, pct_str)

        # Fill CPRF columns (same structure as class block, with coverage included).
        cprf_start_col = class_start_col + len(self.ordered_classes) + 1
        for offset, cls in enumerate(self.ordered_classes):
            metrics = class_avg.get(cls)
            col = cprf_start_col + offset
            if not metrics:
                continue
            self._fill_metric_cell_cprf(sheet, 2, col, metrics, pct_str)

        if avg_metrics:
            cprf_overall_col = cprf_start_col + len(self.ordered_classes)
            self._fill_metric_cell_cprf(sheet, 2, cprf_overall_col, avg_metrics, pct_str)

    def _fill_metric_cell(self, sheet, row: int, col: int, metrics: Dict, pct_str) -> None:
        """Write one `P/R/F1` metric cell in percentage form."""
        precision = metrics.get("precision", 0.0)
        recall = metrics.get("recall", 0.0)
        f1 = metrics.get("f1", 0.0)
        sheet.cell(
            row=row,
            column=col,
            value=f"{pct_str(precision)}/{pct_str(recall)}/{pct_str(f1)}",
        )

    def _fill_metric_cell_cprf(self, sheet, row: int, col: int, metrics: Dict, pct_str) -> None:
        """Write one `Coverage/P/R/F1` metric cell in percentage form."""
        coverage = metrics.get("coverage", 0.0)
        precision = metrics.get("precision", 0.0)
        recall = metrics.get("recall", 0.0)
        f1 = metrics.get("f1", 0.0)
        sheet.cell(
            row=row,
            column=col,
            value=f"{pct_str(coverage)}/{pct_str(precision)}/{pct_str(recall)}/{pct_str(f1)}",
        )

    def _load_dataset(self) -> Dict[str, dict]:
        """Load dataset records from `.json` or `.jsonl` into an index-keyed map."""
        with self.dataset_path.open("r", encoding="utf-8") as f:
            if self.dataset_path.suffix == ".jsonl":
                data = [json.loads(line) for line in f if line.strip()]
            else:
                data = json.load(f)
        return {record["index"]: record for record in data}

    def _parse_gold_checklist(self, record: dict) -> Dict[str, dict]:
        """
        Parse the gold checklist from a dataset record.

        Returns:
            {item_id: {"content": str, "pass": bool, "class": str}}
        """
        gold_items: Dict[str, dict] = {}
        for item in record.get("checklist", []):
            gold_id = str(item["id"])
            gold_items[gold_id] = {
                "content": item["content"],
                "pass": bool(item["pass"]),
                "class": item.get("class"),
            }
        return gold_items
    
    def _parse_pred_checklist(self, result_path: Path) -> Dict[str, dict]:
        """
        Parse predicted checklist items from `result_extracted.md`.

        Primary format (what the prompt asks for):
            - [x] item_id: description     ([x] = pass, [ ] = fail)

        Robust fallback: agents sometimes deviate to a header + status-marker
        layout, e.g. `### FT-01: ...` followed by `**PASS**` / `**Status: FAIL**`.
        Previously such output parsed to ZERO items and the record silently fell
        back to an all-pass checklist (fabricating recall=0). We now recover it.
        """
        text = result_path.read_text(encoding="utf-8")
        return self._parse_pred_items(text)

    def _parse_pred_items(self, text: str) -> Dict[str, dict]:
        if getattr(self, "canonicalize", True):
            text = normalize_to_canonical(text)
        lines = text.splitlines()
        pred_items: Dict[str, dict] = {}

        # 1) Canonical checkbox format.
        cb = re.compile(r"^- \[\s*([xX ])\s*\]\s*(?:\*\*)?([A-Za-z0-9_-]+)(?:\*\*)?:\s*(.+)$")
        for line in lines:
            m = cb.match(line.strip())
            if m:
                pred_items[m.group(2).strip()] = {
                    "content": m.group(3).strip(),
                    "pass": m.group(1).lower() == "x",
                }
        if pred_items:
            return pred_items

        # 2) Fallback: '### <ID> ...' header + nearest following PASS/FAIL marker.
        hdr = re.compile(r"^#{2,4}\s*(?:\*\*)?([A-Z]{2,3}-\d+)(?:\*\*)?\s*[:·–—\-]?\s*(.*)$")
        _dedicated_status = re.compile(
            r"^\*{0,2}\s*(?:(?:status|result)\s*:\s*)?(pass|fail)\s*\*{0,2}$",
            re.IGNORECASE,
        )
        cur = None
        seen_status: Dict[str, bool] = {}
        for line in lines:
            s = line.strip()
            hm = hdr.match(s)
            if hm:
                cur = hm.group(1)
                if re.match(r"BUG-?\d+$", cur, re.IGNORECASE):
                    cur = None
                    continue
                pred_items[cur] = {"content": hm.group(2).strip(), "pass": True}
                seen_status[cur] = False
                continue
            if cur and not seen_status.get(cur):
                sm = re.fullmatch(
                    r"\*{0,2}\s*(?:(?:status|result)\s*:\s*)?(pass|fail)\s*\*{0,2}",
                    s,
                    re.IGNORECASE,
                )
                if sm:
                    pred_items[cur]["pass"] = sm.group(1).upper() == "PASS"
                    seen_status[cur] = True
        return pred_items

    def _parse_checklist_md(self, checklist_path: Path) -> Dict[str, dict]:
        """
        Parse checklist template items from `checklist.md`.

        Notes:
            `checklist.md` carries item definitions only and no pass/fail signal.
            Every parsed item is therefore treated as pass=True.
        """
        pred_items: Dict[str, dict] = {}
        pattern = re.compile(r"^- \[\s*([xX ])\s*\]\s*(?:\*\*)?([A-Za-z0-9_-]+)(?:\*\*)?:\s*(.+)$")
        
        with checklist_path.open("r", encoding="utf-8") as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    item_id = match.group(2).strip()
                    desc = match.group(3).strip()
                    # checklist.md has no pass/fail labels; use pass=True by design.
                    pred_items[item_id] = {"content": desc, "pass": True}
        return pred_items

    def _format_items_for_prompt(self, items: Dict[str, dict]) -> str:
        """Format checklist items into compact lines for the matching prompt."""
        return "\n".join(f"{k}: {v['content']}" for k, v in items.items())

    def _call_api(self, prompt: str, retry: int = 5) -> Tuple[bool, Optional[str], Optional[dict]]:
        """
        Call the LLM API with retry logic.

        Returns:
            (success, response_text, full_message_dict)
        """
        url = self.api_config.base_url
        headers = {
            "Authorization": f"Bearer {self.api_config.api_key}",
            "Content-Type": "application/json",
        }
        data = {
            "model": self.api_config.model,
            "messages": [
                {
                    "role": "user",
                    "content": [{"type": "text", "text": prompt}],
                }
            ],
            # 'reasoning': {'effort': 'minimal'},
            'reasoning': {'effort': 'none'},
            'temperature': 0.1,
            # Reasoning models (e.g. MiniMax-M3) spend tokens on <think> before
            # the answer, so give enough headroom to also emit the match list.
            "max_tokens": 4096
        }

        for attempt in range(1, retry + 1):
            try:
                # Reasoning models can take >20s; allow generous read timeout.
                response = requests.post(url=url, headers=headers, json=data, timeout=120)
                resp = response.json()
                print(f"attempt: {attempt}, resp: {resp}")

                if response.status_code != 200 or "choices" not in resp:
                    print_red(f"[Attempt {attempt}/{retry}] Request failed: {resp}")
                else:
                    message = resp["choices"][0]["message"]
                    answer = message["content"]
                    return True, answer, message

            except Exception as e:
                print_red(f"[Attempt {attempt}/{retry}] Exception: {e}")

            time.sleep(1)

        return False, None, None

    @staticmethod
    def _clean_match_answer(answer: str) -> str:
        """Isolate the Python list literal from a (possibly reasoning) LLM reply.

        Reasoning models such as MiniMax-M3 wrap output in <think>...</think> and
        may add prose or ```python fences, all of which break ast.literal_eval.
        Strip those and return just the outermost [...] literal.
        """
        if not answer:
            return answer
        text = re.sub(r"<think>.*?</think>", "", answer, flags=re.DOTALL | re.IGNORECASE)
        text = re.sub(r"<think>.*$", "", text, flags=re.DOTALL | re.IGNORECASE)  # truncated think
        text = re.sub(r"```[a-zA-Z]*", "", text).replace("```", "")
        start, end = text.find("["), text.rfind("]")
        if start != -1 and end != -1 and end > start:
            text = text[start:end + 1]
        return text.strip()

    def _get_matches(
        self,
        instruction: str, 
        gold_items: Dict[str, dict], 
        pred_items: Dict[str, dict],
        output_dir: Path, 
        source: str, 
        retry: int = 3,
    ) -> Optional[List[Tuple[str, Optional[str]]]]:
        """
        Match predicted items to gold items via an LLM.

        Returns:
            [(pred_id, gold_id), ...]
            `gold_id=None` denotes an unmatched prediction.
        """
        match_result_file = output_dir / "score_match_ids.json"

        # Number of independent matcher ballots to union (D2). Read via getattr so
        # a dropped/forgotten __init__ assignment degrades to single-ballot
        # behavior instead of raising AttributeError (mirrors the canonicalize scar).
        votes = max(1, int(getattr(self, "match_votes", 1)))

        # Reuse cached matches only when stored source AND stored votes match (D4).
        # Legacy caches lacking 'votes' default to 1, so under K=1 they are reused
        # with zero recompute; any K>1 correctly invalidates them (one-time rematch).
        # If the cache is empty but current predictions are non-empty, force rematch
        # so stale empty outputs do not permanently mask valid data.
        if match_result_file.exists():
            match_result = json.loads(match_result_file.read_text(encoding="utf-8"))
            stored_source = match_result.get("source", "result")
            stored_votes = match_result.get("votes", 1)
            if stored_source == source and stored_votes == votes:
                cached_matches = match_result.get("matches")
                if cached_matches:
                    return cached_matches
                if not pred_items:
                    return cached_matches
                print_red(
                    "Cached matches is empty while pred_items is non-empty; rematching..."
                )

        # Empty predictions => there is nothing to match, so casting ballots is
        # pure waste: each is a real (~60-80s MiniMax) matcher call that can only
        # return []. The result-source empty-pred case is already short-circuited
        # upstream in _process_record (_create_empty_match_bundle); this guards
        # the surviving path — missing result + --use_checklist_fallback +
        # checklist.md parsing to 0 items. Persist an empty, self-consistent
        # artifact (so the empty-cache reuse path above hits next time) and return
        # [] — downstream `if not match_ids` treats [] as empty_match either way.
        if not pred_items:
            self._write_json(match_result_file, {
                "matches": [],
                "detailed_matches": [],
                "source": source,
                "votes": votes,
                "ballots": [],
            })
            return []

        # Build prompt once; reused across every ballot.
        prompt = PROMPT_MATCH_ITEM.substitute(
            instruction=instruction,
            gold_items=self._format_items_for_prompt(gold_items),
            pred_items=self._format_items_for_prompt(pred_items),
        )

        # Cast K ballots. _match_once carries the per-ballot parse-retry budget
        # internally; a None ballot is a failure and is DROPPED (no veto). Only if
        # ALL K ballots fail do we return None (same empty_match path as before).
        ballots: List[List[Tuple[str, Optional[str]]]] = []
        for _ in range(votes):
            ballot = self._match_once(prompt, retry=retry)
            if ballot is not None:
                ballots.append(ballot)

        if not ballots:
            return None

        # Aggregate survivors via the pure union function, preserving predicted
        # order so every predicted id appears exactly once even if some ballots
        # omitted it.
        match_ids = aggregate_ballots(ballots, pred_order=list(pred_items.keys()))

        # Build detailed text-level mapping from the AGGREGATED match_ids (NOT a
        # single ballot) so the artifact stays self-consistent with scoring.
        detailed_matches = self._build_detailed_matches(match_ids, gold_items, pred_items)

        # Persist match artifacts for reproducibility. Top-level 'matches' remains
        # the FINAL aggregated list (external readers depend on it); 'ballots' is
        # audit-only.
        self._write_json(match_result_file, {
            "matches": match_ids,
            "detailed_matches": detailed_matches,
            "source": source,
            "votes": votes,
            "ballots": ballots,
        })

        return match_ids

    def _match_once(
        self,
        prompt: str,
        retry: int = 3,
    ) -> Optional[List[Tuple[str, Optional[str]]]]:
        """Cast ONE matcher ballot: up to `retry` _call_api + parse attempts.

        Returns the parsed list of ``(pred_id, gold_id|None)`` pairs on the first
        attempt that yields a valid Python list literal, else ``None`` (a failed
        ballot, which the caller DROPS without vetoing other ballots).

        Two retry layers stay strictly separate: _call_api's internal HTTP/transport
        retry lives inside _call_api; THIS loop is the parse-retry layer. This
        method never writes the cache.
        """
        for attempt in range(1, retry + 1):
            success, answer, _ = self._call_api(prompt)
            print(f"answer: {answer}")

            if not success:
                continue

            try:
                # Parse LLM output as a Python literal list of pairs.
                parsed = ast.literal_eval(self._clean_match_answer(answer))
            except Exception as e:
                print_red(f"[Attempt {attempt}/{retry}] Invalid match format: {e}")
                continue

            # Reject a non-list reply as a parse failure (retry within budget).
            if not isinstance(parsed, list):
                print_red(
                    f"[Attempt {attempt}/{retry}] Match answer is not a list: {type(parsed)}"
                )
                continue
            return parsed

        return None

    def _build_detailed_matches(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict]
    ) -> List[dict]:
        """Build verbose match records including gold/pred text content."""
        # Convert pair list to `gold_id -> [pred_id, ...]`.
        match_map = defaultdict(list)
        for pred_id, gold_id in match_ids:
            match_map[gold_id].append(pred_id)

        detailed_matches = []
        
        # Emit one block per gold item.
        for gold_id, gold_meta in gold_items.items():
            gold_block = {
                "gold": {
                    "id": gold_id,
                    "text": gold_meta.get("content"),
                },
                "pred": None,
            }
            
            # Attach predictions mapped to this gold item.
            for pred_id in match_map.get(gold_id, []):
                pred_meta = pred_items.get(pred_id, {})
                if gold_block["pred"] is None:
                    gold_block["pred"] = []
                gold_block["pred"].append({
                    "id": pred_id,
                    "text": pred_meta.get("content"),
                })
            
            detailed_matches.append(gold_block)
        
        # Add unmatched predictions as a final block with `gold=None`.
        unmatched_preds = match_map.get(None, [])
        if unmatched_preds:
            gold_block = {"gold": None, "pred": []}
            for pred_id in unmatched_preds:
                pred_meta = pred_items.get(pred_id, {})
                gold_block["pred"].append({
                    "id": pred_id,
                    "text": pred_meta.get("content"),
                })
            detailed_matches.append(gold_block)

        return detailed_matches

    def _compute_coverage(
        self, 
        matches: List[Tuple[str, Optional[str]]], 
        gold_items: Dict[str, dict]
    ) -> float:
        """
        Compute coverage as the fraction of gold items matched at least once.
        """
        matched_gold = {gold_id for _, gold_id in matches or [] if gold_id is not None}
        # print(f"matched_gold: {matched_gold}")
        
        if not gold_items:
            return 0.0
        return len(matched_gold) / len(gold_items)

    def _compute_metrics(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict],
    ) -> Dict[str, float]:
        """
        Compute precision/recall/F1 from matched gold and prediction statuses.

        Evaluation logic:
        - TP: gold indicates a bug and prediction reports failure.
        - FP: gold indicates no bug but prediction reports failure.
        - FN: gold indicates a bug but prediction reports pass.
        - TN: gold indicates no bug and prediction reports pass.
        """
        # Build `gold_id -> [pred_id, ...]` mapping from match pairs.
        gold_to_preds: Dict[str, List[str]] = {}
        for pred_id, gold_id in match_ids or []:
            if gold_id is not None:
                gold_to_preds.setdefault(gold_id, []).append(pred_id)

        tp, fp, fn, tn = 0, 0, 0, 0
        
        for gold_id, gold_meta in gold_items.items():
            pred_ids = gold_to_preds.get(gold_id)
            
            # Case 1: this gold item is not covered by any prediction.
            if pred_ids is None:
                if gold_meta["pass"]:  # No bug in gold; uncovered counts as TN.
                    tn += 1
                else:  # Bug in gold; uncovered means miss (FN).
                    fn += 1
                continue
            
            # Case 2: the gold item is covered; derive predicted pass/fail.
            # Any failing prediction marks the item as "bug found".
            pred_pass = all(pred_items[pred_id]["pass"] for pred_id in pred_ids)
            
            if gold_meta["pass"]:  # Gold says no bug.
                if pred_pass:  # Correct no-bug prediction.
                    tn += 1
                else:  # False positive.
                    fp += 1
            else:  # Gold says bug exists.
                if pred_pass:  # False negative.
                    fn += 1
                else:  # Correctly detected bug.
                    tp += 1
        
        # Derive scalar metrics.
        precision = tp / (tp + fp) if (tp + fp) else 0.0
        recall = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = 2 * precision * recall / (precision + recall) if (precision + recall) else 0.0
        metric = {
            "precision": precision,
            "recall": recall,
            "f1": f1,
        }
        
        return metric, gold_to_preds

    def _write_json(self, path: Path, payload: dict) -> None:
        """Write a JSON file with UTF-8 encoding and stable pretty formatting."""
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def _process_record(
        self,
        record_id: str,
        record: dict,
        output_dir: Path,
    ) -> Optional[Dict[str, dict]]:
        """
        Score a single dataset record and persist per-record artifacts.

        Returns:
            {
                "overall": overall metrics,
                "by_class": class-level metrics,
                "match_ids": match pairs,
                "gold_items": parsed gold checklist
            }
        }
        """
        result_path = output_dir / "result_extracted.md"
        checklist_path = output_dir / "checklist.md"
        
        # Handle missing output files.
        missing_result = not result_path.exists()
        if missing_result and not checklist_path.exists():
            print(f"Skipping {record_id}: missing result_extracted.md and checklist.md")
            self.missing_result_ids.append(record_id)
            return self._create_missing_result_bundle(record, output_dir)

        if missing_result and not self.use_checklist_fallback:
            # Fallback disabled: mark as missing and stop early.
            self.missing_result_ids.append(record_id)
            return self._create_missing_result_bundle(record, output_dir)

        if missing_result:
            print(f"Missing result_extracted.md for {record_id}; using checklist.md for matching/coverage")
            self.missing_result_ids.append(record_id)

        # Parse gold checklist and prediction source.
        gold_items = self._parse_gold_checklist(record)
        if missing_result:
            pred_items = self._parse_checklist_md(checklist_path)
            match_source = "checklist"
        else:
            pred_items = self._parse_pred_checklist(result_path)
            match_source = "result"
            # result_extracted.md exists but parsed to zero items => detection
            # output format mismatch. Do NOT silently fall back to checklist.md
            # (treating all items as pass) — that fabricates recall=0 and disguises
            # a parse failure as a detection miss. Flag it and zero/exclude instead.
            if not pred_items:
                print_red(
                    f"result_extracted.md for {record_id} exists but parsed 0 items "
                    "(detection output format mismatch); flagging parse_error "
                    "(NOT treating as all-pass)."
                )
                self.parse_error_ids.append(record_id)
                return self._create_empty_match_bundle(
                    record=record,
                    output_dir=output_dir,
                    pred_items={},
                    missing_result=False,
                    gold_items=gold_items,
                    parse_error=True,
                )

        # Obtain LLM-based matches.
        match_ids = self._get_matches(
            instruction=record["instruction"],
            gold_items=gold_items,
            pred_items=pred_items,
            output_dir=output_dir,
            source=match_source,
        )
        if not match_ids:
            print_red(f"Empty matches for {record_id}, using zero-score fallback")
            self.empty_match_ids.append(record_id)
            return self._create_empty_match_bundle(
                record=record,
                output_dir=output_dir,
                pred_items=pred_items,
                missing_result=missing_result,
                gold_items=gold_items,
            )

        # Compute overall metrics for this record.
        coverage = self._compute_coverage(match_ids, gold_items)
        metrics, gold_to_preds = self._compute_metrics(match_ids, gold_items, pred_items)
        print(f"gold_to_preds: {gold_to_preds}")
        metrics["num_pred_item"] = len(pred_items)
        metrics["coverage"] = coverage
        if missing_result:
            metrics["missing_result"] = True
        
        # Round floats for stable JSON artifacts.
        rounded_metrics = {
            k: (round(v, 4) if isinstance(v, float) else v) 
            for k, v in metrics.items()
        }

        # Compute class-level metrics (only bug-containing classes are scored).
        class_metrics = self._compute_class_metrics(
            match_ids, gold_items, pred_items
        )

        # Persist per-record score output.
        score_payload = {"overall": rounded_metrics}
        for cls in self.ordered_classes:
            score_payload[cls] = class_metrics.get(cls)
        self._write_json(output_dir / "score.json", score_payload)
        
        print_orange(
            f"Scored {record_id}: "
            f"coverage={coverage:.3f}, "
            f"precision={metrics['precision']:.3f}, "
            f"recall={metrics['recall']:.3f}, "
            f"f1={metrics['f1']:.3f}"
        )
        
        return {
            "overall": rounded_metrics,
            "by_class": class_metrics,
            "match_ids": match_ids,
            "gold_items": gold_items,
        }

    def _create_missing_result_bundle(self, record: dict, output_dir: Path) -> dict:
        """Build and write a zeroed score bundle for missing result files."""
        missing_metrics = {
            "precision": 0.0,
            "recall": 0.0,
            "f1": 0.0,
            "coverage": 0.0,
            "num_pred_item": 0,
            "missing_result": True,
        }
        self._write_json(output_dir / "score.json", missing_metrics)
        
        missing_by_class = {
            cls: {"precision": 0.0, "recall": 0.0, "f1": 0.0, "coverage": 0.0}
            for cls in self.ordered_classes
        }
        
        return {
            "overall": missing_metrics,
            "by_class": missing_by_class,
            "match_ids": None,
            "gold_items": self._parse_gold_checklist(record),
        }

    def _create_empty_match_bundle(
        self,
        record: dict,
        output_dir: Path,
        pred_items: Dict[str, dict],
        missing_result: bool,
        gold_items: Dict[str, dict],
        parse_error: bool = False,
    ) -> dict:
        """Build a zero-score bundle for empty matching / unparseable results.

        This record is included in `overall` but excluded from `overall_no_missing`.
        `parse_error=True` flags a result_extracted.md that exists but yielded no
        parseable items (format mismatch) rather than a genuine empty LLM match.
        """
        metrics = {
            "precision": 0.0,
            "recall": 0.0,
            "f1": 0.0,
            "coverage": 0.0,
            "num_pred_item": len(pred_items),
        }
        if parse_error:
            metrics["parse_error"] = True
        else:
            metrics["empty_match"] = True
        if missing_result:
            metrics["missing_result"] = True

        class_metrics = {
            cls: {"precision": None, "recall": None, "f1": None}
            for cls in self.ordered_classes
        }
        score_payload = {"overall": metrics}
        for cls in self.ordered_classes:
            score_payload[cls] = class_metrics[cls]
        self._write_json(output_dir / "score.json", score_payload)

        return {
            "overall": metrics,
            "by_class": class_metrics,
            "match_ids": [],
            "gold_items": gold_items,
        }

    def _compute_class_metrics(
        self,
        match_ids: List[Tuple[str, Optional[str]]],
        gold_items: Dict[str, dict],
        pred_items: Dict[str, dict]
    ) -> Dict[str, dict]:
        """Compute per-class metrics for classes that contain at least one bug."""
        class_metrics: Dict[str, dict] = {}
        
        for cls in self.ordered_classes:
            # Extract gold items of this class.
            gold_subset = {
                gid: item 
                for gid, item in gold_items.items() 
                if item.get("class") == cls
            }
            
            if not gold_subset:
                continue
            
            # Skip classes where gold has no failing item.
            has_bug = any(not info.get("pass", False) for info in gold_subset.values())
            if not has_bug:
                continue
            
            # Reuse core metric computation on class subset.
            cls_metrics, _ = self._compute_metrics(match_ids, gold_subset, pred_items)
            class_metrics[cls] = {
                k: (round(v, 4) if isinstance(v, float) else v) 
                for k, v in cls_metrics.items()
            }
        
        # Backfill unscored classes with None for schema stability.
        for cls in self.ordered_classes:
            if cls not in class_metrics:
                class_metrics[cls] = {
                    "precision": None, 
                    "recall": None, 
                    "f1": None
                }
        
        return class_metrics


def parse_args():
    """Parse command-line arguments for the scoring pipeline."""
    parser = argparse.ArgumentParser(
        description="Score agent outputs against gold checklist."
    )
    parser.add_argument(
        "--dataset_path", type=str,
        help="Path to the WebProber-Bench dataset JSONL file (each line is a record)."
    )
    parser.add_argument(
        "--output_root", type=str,
        help="Root directory for all generated outputs. Each run will create a versioned subdirectory under this root based on --version."
    )
    parser.add_argument(
        "--version", required=True, type=str,
        help="Version label used to group outputs"
    )
    parser.add_argument(
        "--api_base_url", required=True, type=str,
        help="Base URL for API server"
    )
    parser.add_argument(
        "--api_key", required=True, type=str,
        help="API Key for API server"
    )
    parser.add_argument(
        "--api_model", required=True, type=str,
        help="Model name, e.g., openai/gpt-5-mini"
    )
    parser.add_argument(
        "--use_checklist_fallback", nargs="?", const=True, default=False,
        type=lambda v: str(v).lower() in ("1", "true", "yes", "y", "t"),
        help="Allow using checklist.md to match/compute coverage when result_extracted.md is missing. Can be used as a flag or with explicit True/False.",
    )
    parser.add_argument(
        "--no-canonicalize", action="store_false", dest="canonicalize", default=True,
        help="Disable canonical-form normalization (strip phantom BUG-xx, "
             "convert heading/inline) before matching. Normalization is ON by "
             "default (ablation-proven KEEP); use this only for A/B repro.",
    )
    parser.add_argument(
        "--match_votes", type=int, default=1,
        help="Number of independent matcher ballots to union (tau=1). K=1 "
             "(default) is exact current single-call behavior with zero extra "
             "cost; K>1 unions matches appearing in >=1 ballot to recover "
             "false-negative matches. Coerced to an int >= 1.",
    )
    return parser.parse_args()


def main():
    """Program entry point."""
    args = parse_args()

    # Build API client configuration.
    api_config = APIConfig(
        base_url=args.api_base_url,
        api_key=args.api_key,
        model=args.api_model,
    )

    # Resolve run-specific paths.
    dataset_path = Path(args.dataset_path)
    output_root = Path(args.output_root) / args.version

    # Execute scoring.
    pipeline = ScoringPipeline(
        dataset_path=dataset_path,
        output_root=output_root,
        api_config=api_config,
        version=args.version,
        use_checklist_fallback=args.use_checklist_fallback,
        canonicalize=args.canonicalize,
        match_votes=args.match_votes,
    )
    pipeline.run()


if __name__ == "__main__":
    main()
